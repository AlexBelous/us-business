import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# === SALES LOG sheet ===
ws = wb.active
ws.title = "Sales Log"

accent = "1B4F72"
light_bg = "EBF5FB"
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor=accent)
data_font = Font(name="Calibri", size=10)
currency_fmt = '$#,##0.00'
date_fmt = 'MM/DD/YYYY'
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Title
ws.merge_cells('A1:J1')
ws['A1'] = 'DM TRADE SOLUTIONS LLC — Sales Log 2026'
ws['A1'].font = Font(name="Calibri", size=14, bold=True, color=accent)
ws['A1'].alignment = Alignment(horizontal='left')

ws.merge_cells('A2:J2')
ws['A2'] = 'Period: February 15, 2026 — Present'
ws['A2'].font = Font(name="Calibri", size=10, color="666666")

# Headers
headers = [
    ("A", "Invoice #", 14),
    ("B", "Date", 12),
    ("C", "Customer Name", 25),
    ("D", "Description", 35),
    ("E", "Category", 16),
    ("F", "Qty", 8),
    ("G", "Unit Price", 12),
    ("H", "Total Amount", 14),
    ("I", "Payment Method", 16),
    ("J", "Payment Date", 13),
    ("K", "Delivery Method", 16),
    ("L", "Tracking / Notes", 22),
]

for col_letter, header_text, width in headers:
    cell = ws[f'{col_letter}4']
    cell.value = header_text
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[col_letter].width = width

# Sample data rows (5 empty template rows)
alt_fill = PatternFill("solid", fgColor=light_bg)
for row_num in range(5, 25):
    for col_idx, (col_letter, _, _) in enumerate(headers):
        cell = ws[f'{col_letter}{row_num}']
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if row_num % 2 == 0:
            cell.fill = alt_fill
    # Format specific columns
    ws[f'B{row_num}'].number_format = date_fmt
    ws[f'G{row_num}'].number_format = currency_fmt
    ws[f'H{row_num}'].number_format = currency_fmt
    ws[f'G{row_num}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'H{row_num}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'J{row_num}'].number_format = date_fmt
    # Formula: Total = Qty * Unit Price
    ws[f'H{row_num}'] = f'=IF(F{row_num}<>"",F{row_num}*G{row_num},"")'
    ws[f'H{row_num}'].number_format = currency_fmt

# Category dropdown values (as data validation)
from openpyxl.worksheet.datavalidation import DataValidation
cat_dv = DataValidation(type="list", formula1='"Equipment Sale,Parts Sale,Equipment Selection,Repair Service,Commissioning,Other"', allow_blank=True)
cat_dv.error = "Please select a valid category"
cat_dv.errorTitle = "Invalid Category"
ws.add_data_validation(cat_dv)
cat_dv.add(f'E5:E100')

pay_dv = DataValidation(type="list", formula1='"Zelle,Wire Transfer,Check,Cash,Other"', allow_blank=True)
ws.add_data_validation(pay_dv)
pay_dv.add(f'I5:I100')

del_dv = DataValidation(type="list", formula1='"Customer Pickup,USPS,UPS,FedEx,Hand Delivery,Other"', allow_blank=True)
ws.add_data_validation(del_dv)
del_dv.add(f'K5:K100')

# Totals row
total_row = 25
ws[f'G{total_row}'] = 'TOTAL:'
ws[f'G{total_row}'].font = Font(name="Calibri", size=11, bold=True, color=accent)
ws[f'G{total_row}'].alignment = Alignment(horizontal='right')
ws[f'H{total_row}'] = f'=SUM(H5:H{total_row-1})'
ws[f'H{total_row}'].font = Font(name="Calibri", size=11, bold=True, color=accent)
ws[f'H{total_row}'].number_format = currency_fmt
ws[f'H{total_row}'].alignment = Alignment(horizontal='right')
ws[f'H{total_row}'].border = Border(top=Side(style='double', color=accent), bottom=Side(style='double', color=accent))

# Freeze header
ws.freeze_panes = 'A5'

# Print settings
ws.print_title_rows = '4:4'
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = 'landscape'

# === SUMMARY sheet ===
ws2 = wb.create_sheet("Summary")
ws2['A1'] = 'Sales Summary'
ws2['A1'].font = Font(name="Calibri", size=14, bold=True, color=accent)

summary_headers = ["Category", "# of Sales", "Total Revenue"]
for i, h in enumerate(summary_headers):
    cell = ws2.cell(row=3, column=i+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

categories = ["Equipment Sale", "Parts Sale", "Equipment Selection", "Repair Service", "Commissioning", "Other"]
for i, cat in enumerate(categories):
    row = 4 + i
    ws2[f'A{row}'] = cat
    ws2[f'A{row}'].font = data_font
    ws2[f'A{row}'].border = thin_border
    ws2[f'B{row}'] = f'=COUNTIF(\'Sales Log\'!E:E,A{row})'
    ws2[f'B{row}'].font = data_font
    ws2[f'B{row}'].border = thin_border
    ws2[f'B{row}'].alignment = Alignment(horizontal='center')
    ws2[f'C{row}'] = f'=SUMIF(\'Sales Log\'!E:E,A{row},\'Sales Log\'!H:H)'
    ws2[f'C{row}'].font = data_font
    ws2[f'C{row}'].number_format = currency_fmt
    ws2[f'C{row}'].border = thin_border
    if row % 2 == 0:
        for c in ['A', 'B', 'C']:
            ws2[f'{c}{row}'].fill = alt_fill

total_r = 4 + len(categories)
ws2[f'A{total_r}'] = 'TOTAL'
ws2[f'A{total_r}'].font = Font(name="Calibri", size=11, bold=True, color=accent)
ws2[f'B{total_r}'] = f'=SUM(B4:B{total_r-1})'
ws2[f'B{total_r}'].font = Font(name="Calibri", size=11, bold=True, color=accent)
ws2[f'B{total_r}'].alignment = Alignment(horizontal='center')
ws2[f'C{total_r}'] = f'=SUM(C4:C{total_r-1})'
ws2[f'C{total_r}'].font = Font(name="Calibri", size=11, bold=True, color=accent)
ws2[f'C{total_r}'].number_format = currency_fmt

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 16

output_path = '/home/claude/us-business/sales_log/sales_log_2026.xlsx'
wb.save(output_path)
print(f"Sales log created: {output_path}")
